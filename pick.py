#!/usr/bin/env python3

import pandas as pd
import numpy as np
from urllib.request import urlopen
import bs4
import openpyxl
import re
import yaml
import scipy.stats

import argparse

model_names = {
    'line': 'Line (updated)',
    'linemidweek': 'Line (Midweek)',
    'lineopen': 'Line (opening)',
    'linehow': 'Howell',
    'linemarsee': 'Marsee',
    'linedokter': 'Doktor Entropy',
    'linekeep': 'Keeper',
    'lineatom': 'Atomic Football',
    'lineargh': 'ARGH Power Ratings',
    'linecong': 'Dave Congrove',
    'linebillings': 'Billingsley',
    'linebillings2': 'Billingsley+',
    'linecather': 'Catherwood Ratings',
    'lineburdorf': 'Lee Burdorf',
    'linepiratings': 'Pi-Rate Ratings',
    'linepimean': 'Pi-Ratings Mean',
    'linepibias': 'Pi-Rate Bias',
    'linerwp': 'Laffaye RWP',
    'linemassey': 'Massey Ratings',
    'lineclean': 'Cleanup Hitter',
    'linebmc': 'Brent Craig',
    'linepugh': 'ComPughter Ratings',
    'linedwig': 'DP Dwiggins',
    'linefpi': 'ESPN FPI',
    'linecrunch': 'The Sports Cruncher',
    'linebressler': 'Liam Bressler',
    'linetalis': 'Talisman Red',
    'linesag': 'Sagarin Points',
    'linesagr': 'Sagarin Recent',
    'linesagpred': 'Sagarin Ratings',
    'linesaggm': 'Sagarin Golden Mean',
    'linepfz': 'PerformanZ Ratings',
    'linemoore': 'Moore Power Ratings',
    'linenutshell': 'NutShell Sports',
    'lineelo': 'Beck Elo',
    'lineborn': 'Born Power Index',
    'linefox': 'Stat Fox',
    'linelaz': 'Laz Index',
    'linepig': 'Pigskin Index',
    'linekam': 'Edward Kambour',
    'linepayne': 'Payne Power Ratings',
    'lineash': 'Ashby AccuRatings',
    'linesuper': 'Super List',
    'linekerns': 'Stephen Kerns',
    'lineteamrank': 'TeamRankings.com',
    'linecons': 'Massey Consensus',
    'linedonchess': 'Donchess Inference',
    'lineloud': 'Loudsound.org',
    'linecurry': 'Daniel Curry Index',
    'linedunk': 'Dunkel Index',
    'lineavg': 'System Average',
    'linemedian': 'System Median',
    'lineca': 'Computer Adjusted Line'
}


def main(pred, res, slate, names, model, output):

    pred_df = download_predictions(pred)

    results_df = download_results(res)

    games, teams, gotw, noisy_spreads = parse_slate(slate)

    names_dict, reverse_dict = parse_names(names)

    home_match, away_match = match_predictions(pred_df, teams, reverse_dict)

    spreads = parse_spreads(noisy_spreads, home_match, away_match)

    lines = parse_predictions(pred_df, home_match, away_match, model)

    make_predictions(results_df, names_dict, home_match, away_match,
                     spreads, gotw, lines, model, slate, output)



## Read in the current line and computer rankings from the Internet.
def download_predictions(pred):
    with urlopen(pred) as pred_file:
        pred_df = pd.read_csv(pred_file)
    return pred_df


## Read the current model perfomance from the Internet.
def download_results(res):
    with urlopen(res) as results_file:
        bs = bs4.BeautifulSoup(results_file, "lxml")
        results_df = pd.read_html(str(bs), attrs={'class': 'results_table'}, header=0)
        return results_df[0]


## Read in the current slate.
def parse_slate(slate):
    wb = openpyxl.load_workbook(slate)
    ws = wb.active
    game_cells = ws['A']
    game_regex = re.compile(r'(?:\*\*)?(?:#\d+\s+)?(.*?)\s+(?:vs\.?|@)\s+(?:#\d+\s+)?(.*?)(?:\*\*)?$', re.IGNORECASE)
    games = [c.value for c in game_cells if c.value and game_regex.match(c.value)]

    teams = []
    for g in games:
        matches = game_regex.match(g)
        if matches:
            teams.append((matches.group(1), matches.group(2)))

    gotw = [bool(re.search(r'\*\*', g)) for g in games]

    spread_cells = ws['B']
    spread_text = [c.value for c in spread_cells if c.value and re.match(r'^Enter\s+(?!one)', c.value)]
    spread_regex = re.compile(r'^Enter\s+(.*?)\s+iff.*?(\d+)\s+points', re.IGNORECASE)
    noisy_spreads = []
    for s in spread_text:
        matches = spread_regex.match(s)
        if matches:
            noisy_spreads.append((matches.group(1), int(matches.group(2))))
        else:
            noisy_spreads.append((None, None))

    return games, teams, gotw, noisy_spreads


## Read in a translation table between slate names and prediction names.
def parse_names(names):
    with open(names) as names_file:
        names_dict = yaml.load(names_file)

    names_dict = {key.upper(): val for key, val in names_dict.items()}
    reverse_dict = {val: key for key, val in names_dict.items()}
    return names_dict, reverse_dict

# Make picks.

## Get proper names
def match_predictions(pred_df, teams, reverse_dict):
    pred_df['home'] = pred_df['home'].str.upper().str.replace(r'\.', '')
    pred_df['road'] = pred_df['road'].str.upper().str.replace(r'\.', '')

    home_match = []
    away_match = []

    def find_match(slate_team, home_road):
        if slate_team not in reverse_dict:
            if slate_team in pred_df[home_road]:
                print("'{}' ({} team from slate) not found in names dict, but found in predictions".format(slate_team, home_road))
                return slate_team
            else:
                raise KeyError("'{}' ({} team from slate) not found in names dict!".format(slate_team, home_road))
        else:
            return reverse_dict[slate_team]

    for team_pair in teams:
        away_match.append(find_match(team_pair[0], 'road'))
        home_match.append(find_match(team_pair[1], 'home'))

    return home_match, away_match


## Parse spreads
def parse_spreads(noisy_spreads, home_match, away_match):
    spreads = []
    for s in noisy_spreads:
        if s[0]:
            points = s[1]
            if s[0] in away_match:
                points *= -1
            spreads.append(points)
        else:
            spreads.append(0)
    return spreads


## Parse the lines from the predictions
def parse_predictions(pred_df, home_match, away_match, model):
    assert model in model_names

    lines = []

    for home, away in zip(home_match, away_match):
        home_index = pred_df['home'] == home
        if not home_index.any():
            print("'{}' (home team from slate) not found in predictions!".format(home))
        away_index = pred_df['road'] == away
        if not away_index.any():
            print("'{}' (road team from slate) not found in predictions!".format(away))
        n_found = (away_index | home_index).sum()
        if n_found == 0:
            print("'{}' @ '{}' not found in predictions!".format(away, home))
            lines.append(np.nan)
        elif n_found > 1:
            raise KeyError("'{}' and '{}' are not playing each other this week!".format(away, home))
        else:
            index = (away_index | home_index)
            lines.append(pred_df.loc[index, model].tolist()[0])

    return lines


## Make picks
def make_predictions(results_df, names_dict, home_match, away_match, spreads, gotw, lines, model, slate, output):

    assert model in model_names

    line_results = results_df.loc[results_df['System'] == model_names[model]]
    bias = line_results['Bias'].tolist()[0]
    mse = line_results['Mean Square Error'].tolist()[0]
    std_dev = np.sqrt(mse - bias*bias)

    picks = []
    spread_probs = []
    pred_points = []

    for home, away, gw, spread, line in zip(home_match, away_match, gotw, spreads, lines):
        if np.isnan(line):
            spread_probs.append(np.nan)
            picks.append(np.nan)
            pred_points.append(np.nan)
            continue

        prob = 1 - scipy.stats.norm.cdf(spread, loc=line - bias, scale=std_dev)
        if prob > .5:
            picks.append(names_dict[home])
        else:
            picks.append(names_dict[away])
            prob = 1 - prob
        spread_probs.append(prob)
        if gw:
            prob *= 2
        pred_points.append(prob)

    wb = openpyxl.load_workbook(slate)
    ws = wb.active

    ws.cell(row=1, column=4, value="Probability of Correct Pick").font = openpyxl.styles.Font(bold=True)
    ws.cell(row=1, column=5, value="Predicted Margin").font = openpyxl.styles.Font(bold=True)
    ws.cell(row=1, column=6, value="Notes").font = openpyxl.styles.Font(bold=True)

    for i, (pick, prob, line) in enumerate(zip(picks, spread_probs, lines)):
        ws.cell(row=i+2, column=3, value=pick)
        ws.cell(row=i+2, column=4, value=prob)
        ws.cell(row=i+2, column=5, value=line - bias)

    wb.save(output)


if __name__ == '__main__':

    parser = argparse.ArgumentParser(description="Make your picks for you!")

    parser.add_argument('--slate', '-s', help="This week's slate",
                        required=True)

    parser.add_argument('--output', '-o', help="Output Excel file name",
                        required=True)

    parser.add_argument('--pred', '-p', help="URL of NCAA predictions file",
                        default="http://www.thepredictiontracker.com/ncaapredictions.csv")

    parser.add_argument('--res', '-r', help="URL containing HTML table of model performance and results",
                        default="http://www.thepredictiontracker.com/ncaaresults.php")

    parser.add_argument('--names', '-n', help="File containing translation from slate names to prediction names",
                        default="names.yaml")

    parser.add_argument('--model', '-m', help="Model name to use (typically begins with 'line')",
                        default='line')

    args = parser.parse_args()

    main(**vars(args))
