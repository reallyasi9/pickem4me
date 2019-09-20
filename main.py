import google.cloud.exceptions
from google.cloud import firestore
from google.cloud import error_reporting
from google.cloud import storage
from absl import logging
from absl import flags
from absl import app
import scipy.stats
import sys
import openpyxl
import io

FLAGS = flags.FLAGS

flags.DEFINE_string("slate", None, "Firestore path to slate document")
flags.DEFINE_string("tracker", None, "Firestore path to prediction tracker document")
flags.DEFINE_string(
    "model",
    None,
    "Name of model to use (starts with 'line', defaults to best model for each of the pick categories)",
)
flags.DEFINE_string("bucket", "b1g-pick-em-slates", "Output Cloud Storage bucket name")
flags.DEFINE_boolean('debug', False, 'Produces debugging output')

FS = firestore.Client()
ER = error_reporting.Client()
CS = storage.Client()


def get_best_models(tracker_ref):
    best_picker = next(
        tracker_ref.collection("modelperformance")
        .order_by("pct_correct", direction="DESCENDING")
        .limit(1)
        .stream()
    )
    best_spread = next(
        tracker_ref.collection("modelperformance")
        .order_by("std_dev", direction="DESCENDING")
        .limit(1)
        .stream()
    )
    best_dogger = next(
        tracker_ref.collection("modelperformance")
        .order_by("pct_against_spread", direction="DESCENDING")
        .limit(1)
        .stream()
    )

    return {
        "picker": _get_model_parameters(best_picker),
        "spread": _get_model_parameters(best_spread),
        "dogger": _get_model_parameters(best_dogger),
    }


def _get_model_parameters(model_doc):

    md = model_doc.to_dict()
    if "model_id" not in md:
        logging.fatal("model %s has not been checked--run `tpt-check` and try again", model_doc.id)

    return md


def _model_to_dist(model):
    return scipy.stats.norm(-model["bias"], model["std_dev"])


def pickem(argv):
    
    if FLAGS.debug:
        logging.set_verbosity(logging.DEBUG)

    # Read slate
    slate_path = FLAGS.slate
    if slate_path is None:
        slates_ref = (
            FS.collection("slates")
            .order_by("timestamp", direction="DESCENDING")
            .limit(1)
        )
        slate = next(slates_ref.stream())
        slate_path = slate.reference.path

    logging.debug("reading slate %s", slate_path)
    slate_ref = FS.document(slate_path)

    # Read predictions
    tracker_path = FLAGS.tracker
    if tracker_path is None:
        trackers_ref = (
            FS.collection("prediction_tracker")
            .order_by("timestamp", direction="DESCENDING")
            .limit(1)
        )
        tracker = next(trackers_ref.stream())
        tracker_path = tracker.reference.path

    logging.debug("reading prediction tracker %s", tracker_path)
    tracker_ref = FS.document(tracker_path)

    # Get the model parameters
    use_model = FLAGS.model
    if use_model is None or use_model == "best":
        use_models = get_best_models(tracker_ref)
    else:
        model_doc = tracker_ref.collection("modelperformance").document(use_model).get()
        params = _get_model_parameters(model_doc)
        use_models = {
            "picker": params,
            "spread": params,
            "dogger": params,
        }  # all the same for picks, spreads, and dogs

    logging.debug("using trackers %s", str(use_models))
    dists = {key: _model_to_dist(value) for key, value in use_models.items()}

    # make an output excel spreadsheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "picks"

    games_ref = slate_ref.collection("games")
    for game in games_ref.stream():
        gd = game.to_dict()

        # Lookup game in predictions
        switch = False
        try:
            # traditionally, home team is listed second
            prediction = next(tracker_ref.collection("predictions").where("home_id", "==", gd["team2_id"]).where("road_id", "==", gd["team1_id"]).stream())
        except StopIteration:
            # maybe home/away is backward?
            try:
                prediction = next(tracker_ref.collection("predictions").where("road_id", "==", gd["team2_id"]).where("home_id", "==", gd["team1_id"]).stream())
                switch = True
            except StopIteration:
                warn = f"prediction for game with team ids `{gd['team1_id']}` and `{gd['team2_id']}` not found: skipping"
                logging.warning(warn)
                ER.report(warn)
                continue

        if switch and not gd["neutral"]:
            warn = f"game listed as `{gd['team1']}` @ `{gd['team2']}` has home team misidentified in slate"
            logging.warning(warn)
            ER.report(warn)

        # adjustment for noisy spread
        adjustment = 0.
        version = "picker"
        if gd["noisy_spread"] is not None:
            adjustment += gd["noisy_spread"]  # pos favors team2 (traditionally the home team)
            version = "spread"

        # pick!
        pdict = prediction.to_dict()
        model_id = use_models[version]["model_id"]
        spread = pdict[model_id]

        if spread is None:
            warn = f"game listed as `{gd['team1']}` @ `{gd['team2']}` is not predicted by {version} pick model `{model_id}`: not picking"
            logging.warning(warn)
            ER.report(warn)
            continue

        adjusted_spread = spread - adjustment
        if switch:
            adjusted_spread = -adjusted_spread
        phome = dists[version].cdf(adjusted_spread)
        
        prob = phome
        if phome < 0.5:
            pick = gd['team1']
            prob = 1-phome
        else:
            pick = gd['team2']

        logging.info("%s @ %s (by %d): picking %s with p=%f from model %s (predicting %f)", gd['team1'], gd['team2'], adjustment, pick, prob, model_id, spread)
        pickem_game(ws, gd, pick, prob, model_id, spread)

    # superdog picks
    dog_expected_points = []
    dogs_ref = slate_ref.collection("superdogs")
    for dog in dogs_ref.stream():
        dd = dog.to_dict()

        # Lookup game in predictions
        switch = False
        try:
            # arbitrarily pick the underdog to be home
            prediction = next(tracker_ref.collection("predictions").where("home_id", "==", dd["underdog_id"]).where("road_id", "==", dd["overdog_id"]).stream())
        except StopIteration:
            # maybe home/away is backward?
            try:
                prediction = next(tracker_ref.collection("predictions").where("road_id", "==", dd["underdog_id"]).where("home_id", "==", dd["overdog_id"]).stream())
                switch = True
            except StopIteration:
                warn = f"prediction for superdog game with team ids `{dd['underdog_id']}` and `{dd['overdog_id']}` not found: skipping"
                logging.warning(warn)
                ER.report(warn)
                continue

        # pick!
        pdict = prediction.to_dict()
        model_id = use_models["dogger"]["model_id"]
        spread = pdict[model_id]

        if spread is None:
            warn = f"superdog game listed as `{dd['underdog']}` beats `{dd['overdog']}` is not predicted by pick model `{model_id}`: not picking"
            logging.warning(warn)
            ER.report(warn)
            continue

        adjusted_spread = spread
        if switch:
            adjusted_spread = -adjusted_spread

        phome = dists["dogger"].cdf(adjusted_spread)
        
        prob = phome
        expected = prob * dd["value"]

        logging.info("%s over %s: p=%f and expected value %f from model %s (predicting %f)", dd['underdog'], dd['overdog'], prob, expected, model_id, adjusted_spread)
        dd["expected"] = expected
        dog_expected_points.append(dd)
        pickem_dog(ws, dd, prob, expected, model_id, adjusted_spread)
    
    best_dogs = sorted(dog_expected_points, key=lambda x: x["expected"], reverse=True)
    pickem_best_dog(ws, best_dogs[0])

    # picks are in!
    output = io.BytesIO()
    wb.save(output)

    # original file name is stored in the slate
    slate_doc = slate_ref.get()
    slate_dict = slate_doc.to_dict()
    outname = "picks_" + slate_dict["file"]
    blob = CS.bucket(FLAGS.bucket).blob("picks/" + outname)
    blob.upload_from_file(output, rewind=True)



def format_game(gd):
    """Format a game for printing"""
    gotw = ""
    if gd["gotw"]:
        gotw = "**"
    r1 = ""
    if gd["rank1"]:
        r1 = f"#{gd['rank1']}"
    t1 = gd["team1"]
    vs = "@"
    if gd["neutral_site"]:
        vs = "vs"
    r2 = ""
    if gd["rank2"]:
        r2 = f"#{gd['rank2']}"
    t2 = gd["team2"]
    return " ".join([gotw, r1, t1, vs, r2, t2, gotw]).replace("  ", " ").replace("** ", "**").replace(" **", "**")


def format_noisy_spread(gd):
    """Format a noisy spread string"""
    noisy_spread = gd["noisy_spread"]
    if not noisy_spread:
        return "Enter name of predicted winner"
    favored = gd["team2"]
    if noisy_spread < 0:
        favored = gd["team1"]
        noisy_spread = -noisy_spread
    return f"Enter {favored} iff you predict {favored} wins by at least {int(noisy_spread)} points."


def pickem_game(ws, gd, pick, prob, model_id, spread):
    """Writes game and pick stats to worksheet ws"""
    row = gd["pick_row"]
    col = gd["pick_col"]
    game_text = format_game(gd)
    ws.cell(row, col-2, game_text)
    ns_text = format_noisy_spread(gd)
    ws.cell(row, col-1, ns_text)
    ws.cell(row, col, pick)
    ws.cell(row, col+1, prob)
    ws.cell(row, col+2, model_id)
    ws.cell(row, col+3, spread)


def format_dog(dd):
    """Format a superdog game"""
    under = dd["underdog"]
    over = dd["overdog"]
    value = int(dd["value"])
    return f"{under} over {over} ({value} points, if correct)"


def pickem_dog(ws, dd, prob, expected, model_id, adjusted_spread):
    """Writes superdog game to worksheet ws"""
    row = dd["pick_row"]
    col = dd["pick_col"]
    dog_text = format_dog(dd)
    ws.cell(row, col-1, dog_text)
    ws.cell(row, col+1, prob)
    ws.cell(row, col+2, expected)
    ws.cell(row, col+3, model_id)
    ws.cell(row, col+4, adjusted_spread)


def pickem_best_dog(ws, dd):
    """Marks the best superdog on worksheet ws"""
    row = dd["pick_row"]
    col = dd["pick_col"]
    ws.cell(row, col, dd["underdog"])


if __name__ == "__main__":

    logging.info("Function triggered with arguments: %s", str(sys.argv[1:]))
    app.run(pickem)
